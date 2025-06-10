import sys
import os
import re
import openpyxl
from pathlib import Path

# ===== flatten_split_keys =====
def flatten_split_keys(value, parent_key=""):
    """
    任意のネストdict/listを 'foo.bar[0].baz' のような連結キー:値 のdictに展開する。
    """
    items = {}
    if isinstance(value, dict):
        for k, v in value.items():
            new_key = f"{parent_key}.{k}" if parent_key else k
            items.update(flatten_split_keys(v, new_key))
    elif isinstance(value, list):
        for i, elem in enumerate(value):
            new_key = f"{parent_key}[{i}]" if parent_key else f"[{i}]"
            items.update(flatten_split_keys(elem, new_key))
    else:
        items[parent_key] = value
    return items

# ===== set_nested_dict_from_concat_key =====
def set_nested_dict_from_concat_key(data, keys, value):
    """
    連結キー 'foo.bar[0].baz' 形式のリストを元に、dict/list構造を再帰的に構築して値をセットする。
    """
    key = keys[0]
    m = re.match(r'(\w+)\[(\d+)\]', key)
    if m:
        k, idx = m.group(1), int(m.group(2))
        if k not in data:
            data[k] = []
        while len(data[k]) <= idx:
            data[k].append({})
        if len(keys) == 1:
            data[k][idx] = value
        else:
            if not isinstance(data[k][idx], dict):
                data[k][idx] = {}
            set_nested_dict_from_concat_key(data[k][idx], keys[1:], value)
    else:
        if len(keys) == 1:
            data[key] = value
        else:
            if key not in data or not isinstance(data[key], dict):
                data[key] = {}
            set_nested_dict_from_concat_key(data[key], keys[1:], value)

# ===== format_hcl_value =====
def format_hcl_value(name, val, indent, eqpad="", is_map=False):
    """
    HCL値のフォーマット。空値は出さず、Excelで入ってきた値を忠実に再現。
    - 0は0、0.0は0.0でそのまま
    """
    ind = '  ' * indent
    eqpad = eqpad or ''
    if val is None or val == "" or (isinstance(val, list) and len(val) == 0):
        return ""
    if isinstance(val, str):
        m = re.fullmatch(r"(\$\{|\{\$)([^}]+)\}", val)
        if m:
            expr = m.group(2)
            return f"{ind}{name}{eqpad} = {expr}\n"
        else:
            return f"{ind}{name}{eqpad} = \"{val}\"\n"
    elif isinstance(val, bool):
        return f"{ind}{name}{eqpad} = {'true' if val else 'false'}\n"
    elif isinstance(val, int):
        return f"{ind}{name}{eqpad} = {val}\n"
    elif isinstance(val, float):
        # パラシがfloatならそのまま出す
        return f"{ind}{name}{eqpad} = {val}\n"
    elif isinstance(val, list):
        arr = []
        for x in val:
            if x is None or x == "":
                continue
            if isinstance(x, str):
                m = re.fullmatch(r"(\$\{|\{\$)([^}]+)\}", x)
                if m:
                    arr.append(m.group(2))
                else:
                    arr.append(f"\"{x}\"")
            elif isinstance(x, bool):
                arr.append("true" if x else "false")
            elif isinstance(x, int):
                arr.append(str(x))
            elif isinstance(x, float):
                arr.append(str(x))
            else:
                arr.append(str(x))
        if not arr:
            return f"{ind}{name}{eqpad} = []\n"
        return f"{ind}{name}{eqpad} = [{', '.join(arr)}]\n"
    else:
        return f"{ind}{name}{eqpad} = {val}\n"

# ===== dict_to_hcl_block =====
def dict_to_hcl_block(name, val, indent=0):
    """
    ネストdict/listをHCLブロックとして再帰的に出力。
    - 空値/None属性は出力しない。
    """
    ind = '  ' * indent
    if isinstance(val, dict):
        if all(not isinstance(v, (dict, list)) for v in val.values()):
            keys = list(val.keys())
            maxlen = max(len(k) for k in keys) if keys else 0
            hcl = f"{ind}{name} = {{\n"
            for k in keys:
                v = val[k]
                eqpad = ' ' * (maxlen - len(k))
                out = format_hcl_value(k, v, indent+1, eqpad, is_map=True)
                if out:
                    hcl += out
            hcl += f"{ind}}}\n"
            return hcl
        kvs = []
        blocks = []
        for k, v in val.items():
            if isinstance(v, dict) or (isinstance(v, list) and v and isinstance(v[0], dict)):
                blocks.append((k, v))
            else:
                kvs.append((k, v))
        maxlen = max((len(k) for k, _ in kvs), default=0)
        hcl = f"{ind}{name} {{\n"
        for k, v in kvs:
            eqpad = ' ' * (maxlen - len(k))
            out = format_hcl_value(k, v, indent+1, eqpad)
            if out:
                hcl += out
        for k, v in blocks:
            hcl += dict_to_hcl_block(k, v, indent+1)
        hcl += f"{ind}}}\n"
        return hcl
    elif isinstance(val, list):
        # list of dict: block反復
        if all(isinstance(i, dict) for i in val):
            blocks = ""
            for v in val:
                blocks += dict_to_hcl_block(name, v, indent)
            return blocks
        # list of primitive: 空リストは [] のみ
        elif len(val) == 0:
            return f"{ind}{name} = []\n"
        else:
            arr = []
            for x in val:
                if x is None or x == "":
                    continue
                if isinstance(x, str):
                    m = re.fullmatch(r"(\$\{|\{\$)([^}]+)\}", x)
                    if m:
                        arr.append(m.group(2))
                    else:
                        arr.append(f"\"{x}\"")
                else:
                    arr.append(str(x))
            if not arr:
                return f"{ind}{name} = []\n"
            return f"{ind}{name} = [{', '.join(arr)}]\n"
    else:
        return format_hcl_value(name, val, indent)

# ===== dict_to_resource_hcl =====
def dict_to_resource_hcl(d):
    """
    辞書形式データからHCLリソース記法を出力する（resource_type, res_name, 属性dict構造）。
    """
    hcl = ""
    for res_type, res_objs in d.items():
        for res_name, content in res_objs.items():
            hcl += f'resource "{res_type}" "{res_name}" {{\n'
            keys = list(content.keys())
            maxlen = max((len(k) for k in keys), default=0)
            for k in keys:
                v = content[k]
                eqpad = ' ' * (maxlen - len(k))
                if isinstance(v, dict) or (isinstance(v, list) and v and isinstance(v[0], dict)):
                    continue
                out = format_hcl_value(k, v, 1, eqpad)
                if out:
                    hcl += out
            for k, v in content.items():
                if isinstance(v, dict) or (isinstance(v, list) and v and isinstance(v[0], dict)):
                    hcl += dict_to_hcl_block(k, v, 1)
            hcl += '}\n\n'
    return hcl

# ===== find_header_row =====
def find_header_row(ws, *header_words):
    """
    wsから、指定されたヘッダワード群が揃っている行番号・各列indexを返す。
    """
    for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if all(word in row for word in header_words):
            indices = [row.index(word) for word in header_words]
            return idx, *indices
    return None

# ===== export_hcl_to_excel =====
def export_hcl_to_excel(tf_path, output_excel):
    """
    Terraform HCLファイルをパースし、リソース属性情報をExcelファイルへ出力する。
    - resource_type/res_name/parent_key/child_key/値/連結キー の形式で書き出し
    """
    import hcl2
    with Path(tf_path).open("r") as tf_file:
        parsed = hcl2.load(tf_file)
    terraform_data = {}
    for block in parsed.get("resource", []):
        for resource_type, resource_instances in block.items():
            for res_name, res_body in resource_instances.items():
                flat_attrs = flatten_split_keys(res_body)
                for full_key, val in flat_attrs.items():
                    if "." in full_key:
                        parent_key, child_key = full_key.rsplit(".", 1)
                    else:
                        parent_key, child_key = "", full_key
                    terraform_data[(resource_type, res_name, parent_key, child_key)] = val

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "tf_key_list"
    ws.cell(row=1, column=1, value="resource_type")
    ws.cell(row=1, column=2, value="res_name")
    ws.cell(row=1, column=3, value="parent_key")
    ws.cell(row=1, column=4, value="child_key")
    ws.cell(row=1, column=5, value="値")
    ws.cell(row=1, column=6, value="連結キー")

    for idx, ((resource_type, res_name, parent_key, child_key), value) in enumerate(terraform_data.items(), start=2):
        concat_key = ".".join([x for x in [resource_type, res_name, parent_key, child_key] if x])
        ws.cell(row=idx, column=1, value=resource_type)
        ws.cell(row=idx, column=2, value=res_name)
        ws.cell(row=idx, column=3, value=parent_key)
        ws.cell(row=idx, column=4, value=child_key)
        ws.cell(row=idx, column=5, value=str(value))
        ws.cell(row=idx, column=6, value=concat_key)

    wb.save(output_excel)
    print(f"出力完了：{output_excel}")

# ===== export_excel_to_hcl =====
def export_excel_to_hcl(input_excel, output_hcl,
                        concat_key_header="連結キー",
                        tf_col_header="tf設定値"):
    """
    ExcelパラシートからHCL（main.tf）ファイルを出力する。
    - 連結キーからresource_type, res_name, 属性パスを分割
    - tf_col_headerでテンプレ値を抽出
    """
    wb = openpyxl.load_workbook(input_excel, data_only=True)
    tf_data = {}
    for ws in wb.worksheets:
        res = find_header_row(ws, concat_key_header, tf_col_header)
        if not res:
            continue
        header_row_idx, concat_idx, tf_idx = res
        for row in ws.iter_rows(min_row=header_row_idx+1, values_only=True):
            concat_key = row[concat_idx]
            tf_val = row[tf_idx]
            if not concat_key or tf_val in (None, ""):
                continue
            parts = concat_key.split(".")
            if len(parts) < 3:
                print(f"連結キー不正: {concat_key}（スキップ）")
                continue
            resource_type, res_name = parts[0], parts[1]
            attr_path = parts[2:]
            tf_data.setdefault(resource_type, {})
            tf_data[resource_type].setdefault(res_name, {})
            set_nested_dict_from_concat_key(tf_data[resource_type][res_name], attr_path, tf_val)
    hcl = dict_to_resource_hcl(tf_data)
    with open(output_hcl, "w", encoding="utf-8") as f:
        f.write(hcl)
    print(f"HCL出力完了：{output_hcl}")

# ===== reflect_tf_to_excel =====
def reflect_tf_to_excel(tf_path, input_excel, output_excel):
    """
    Terraform HCLファイルの値を、Excel（全シート）の対応行に反映し、保存する。
    - 連結キー突合。値のみ上書き
    """
    import hcl2
    with Path(tf_path).open("r") as tf_file:
        parsed = hcl2.load(tf_file)
    terraform_data = {}
    for block in parsed.get("resource", []):
        for resource_type, resource_instances in block.items():
            for res_name, res_body in resource_instances.items():
                flat_attrs = flatten_split_keys(res_body)
                for full_key, val in flat_attrs.items():
                    if "." in full_key:
                        parent_key, child_key = full_key.rsplit(".", 1)
                    else:
                        parent_key, child_key = "", full_key
                    concat_key = ".".join([x for x in [resource_type, res_name, parent_key, child_key] if x])
                    terraform_data[concat_key] = val
    wb = openpyxl.load_workbook(input_excel)
    for ws in wb.worksheets:
        res = find_header_row(ws, "連結キー", "値")
        if not res:
            continue
        header_row_idx, concat_idx, val_idx = res
        for row in ws.iter_rows(min_row=header_row_idx+1):
            concat_key = row[concat_idx].value
            if concat_key and concat_key in terraform_data:
                row[val_idx].value = terraform_data[concat_key]
    wb.save(output_excel)
    print(f"Excel反映完了：{output_excel}")

def extract_vars_from_dict(d):
    """
    ネストdictから ${var.xxx} または var.xxx を全て検出してセットで返す
    """
    import re
    vars_found = set()
    def scan(val):
        if isinstance(val, dict):
            for v in val.values():
                scan(v)
        elif isinstance(val, list):
            for v in val:
                scan(v)
        elif isinstance(val, str):
            # ${var.xxx} または {$var.xxx} または var.xxx どちらも検出
            for m in re.finditer(r"(?:\$\{var\.([a-zA-Z0-9_]+)\}|\{\$var\.([a-zA-Z0-9_]+)\}|var\.([a-zA-Z0-9_]+))", val):
                g = m.group(1) or m.group(2) or m.group(3)
                if g:
                    vars_found.add(g)
    scan(d)
    return vars_found

# ===== export_excel_to_tf_and_tfvars =====
def export_excel_to_tf_and_tfvars(input_excel, output_dir,
                                 concat_key_header="連結キー",
                                 tf_col_header="tf設定値",
                                 tfvars_col_header="tfvars設定値"):
    """
    Excelパラシートから、各シートごとにmain.tf（HCL）、tfvars、variables.tfを個別出力する。
    - tfvarsはvar名のみ抽出、シートごとに1ファイル。=の位置もそろえる
    """
    wb = openpyxl.load_workbook(input_excel, data_only=True)
    os.makedirs(output_dir, exist_ok=True)
    for ws in wb.worksheets:
        res = find_header_row(ws, concat_key_header, tf_col_header, tfvars_col_header)
        if not res:
            print(f"シート「{ws.title}」で列が見つからずスキップ。")
            continue
        header_row_idx, concat_idx, tf_idx, tfvars_idx = res

        tf_data = {}
        tfvars_data = {}

        for row in ws.iter_rows(min_row=header_row_idx+1, values_only=True):
            concat_key = row[concat_idx]
            tf_val = row[tf_idx]
            tfvars_val = row[tfvars_idx]
            if not concat_key:
                continue
            parts = concat_key.split(".")
            if len(parts) < 3:
                print(f"連結キー不正: {concat_key}（スキップ）")
                continue
            resource_type, res_name = parts[0], parts[1]
            attr_path = parts[2:]
            # main.tf用
            if tf_val not in (None, ""):
                tf_data.setdefault(resource_type, {})
                tf_data[resource_type].setdefault(res_name, {})
                set_nested_dict_from_concat_key(tf_data[resource_type][res_name], attr_path, tf_val)
            # tfvars用（var名のみ抽出）
            if tfvars_val not in (None, "") and tf_val and isinstance(tf_val, str):
                m = re.search(r"\{\$var\.([a-zA-Z0-9_]+)\}", tf_val)
                if m:
                    varname = m.group(1)
                    tfvars_data[varname] = tfvars_val

        # main.tf, variables.tf
        if tf_data:
            for resource_type, res_objs in tf_data.items():
                tf_file = os.path.join(output_dir, f"{resource_type}.tf")
                hcl = dict_to_resource_hcl({resource_type: res_objs})
                with open(tf_file, "w", encoding="utf-8") as f:
                    f.write(hcl)
                used_vars = set()
                for content in res_objs.values():
                    used_vars |= extract_vars_from_dict(content)
                if used_vars:
                    vars_file = os.path.join(output_dir, f"{resource_type}.variables.tf")
                    with open(vars_file, "w", encoding="utf-8") as vf:
                        for var in sorted(used_vars):
                            vf.write(f'variable "{var}" {{\n  default = "undefined"\n}}\n\n')

        # tfvars出力（=位置パディング。シートごとに1ファイルでOK）
        if tfvars_data:
            for resource_type in set(k.split(".")[0] for k in tf_data.keys()):
                tfvars_file = os.path.join(output_dir, f"{resource_type}.tfvars")
                tfvars_keys = list(tfvars_data.keys())
                maxlen = max(len(k) for k in tfvars_keys)
                with open(tfvars_file, "w", encoding="utf-8") as f:
                    for k in tfvars_keys:
                        eqpad = " " * (maxlen - len(k))
                        v = tfvars_data[k]
                        f.write(f"{k}{eqpad} = \"{v}\"\n")
    print(f"出力完了：{output_dir}/*.tf, *.tfvars, *.variables.tf")

# ===== usage =====
def usage():
    """
    使い方ヘルプを表示して終了。
    """
    print("python xlsx2tf.py export_hcl_to_excel <tf_path> <output_excel>")
    print("python xlsx2tf.py export_excel_to_hcl <input_excel> <output_hcl> [連結キー列名] [tf設定値列名]")
    print("python xlsx2tf.py reflect_tf_to_excel <tf_path> <input_excel> <output_excel>")
    print("python xlsx2tf.py export_excel_to_tf_and_tfvars <input_excel> <output_dir> [連結キー列名] [tf設定値列名] [tfvars設定値列名]")
    sys.exit(1)

# ===== CLI =====
funcs = {
    "export_hcl_to_excel": export_hcl_to_excel,
    "export_excel_to_hcl": export_excel_to_hcl,
    "reflect_tf_to_excel": reflect_tf_to_excel,
    "export_excel_to_tf_and_tfvars": export_excel_to_tf_and_tfvars,
}

if __name__ == "__main__":
    if len(sys.argv) < 2:
        usage()
    func_name = sys.argv[1]
    args = sys.argv[2:]
    if func_name in funcs:
        try:
            funcs[func_name](*args)
        except Exception as e:
            print(f"エラー: {e}")
            usage()
    else:
        usage()
